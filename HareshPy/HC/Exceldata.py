import openpyxl
path= "/TestData/Credentials.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
print(sheet.max_row)
print(sheet.max_column)
print(sheet.cell(3,2).value)
